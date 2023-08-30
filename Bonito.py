import openpyxl
import telebot
from telebot import types
import Const
from openpyxl import workbook
import random
bot = telebot.TeleBot('6574011523:AAFsGvp18DE5BSMoOdfW6AShtO33cfvhS6k')

const = Const
user_inputs = {}
admin_inputs = []
print('Hello world')

@bot.message_handler(commands=['admin'])
def admin(message):
    user_inputs[message.chat.id] = {}
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Добавить товар", callback_data='add_item'))
    markup.add(types.InlineKeyboardButton("Удалить товар", callback_data='delete_item'))
    bot.send_message(message.chat.id, 'Приветствую тебя хозяин', reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == 'add_item')
def add_item_admin(call):
    message = call.message
    key = '_admin'
    print('add fun')
    bot.send_message(message.chat.id, 'Введите размер одежды')
    bot.register_next_step_handler(message, admin_size, key)


@bot.callback_query_handler(func=lambda call: call.data == 'delete_item')
def delete_item_admin(call):
    message = call.message
    key = '_delete'
    create_button(message, key)


def admin_size(message, key):
    size = message.text.strip()
    user_inputs[message.chat.id]['size'] = size
    admin_inputs.append(size)
    bot.send_message(message.chat.id, 'Введите цену товара')
    bot.register_next_step_handler(message, admin_price, key)


def admin_price(message, key):
    price = message.text.strip()
    user_inputs[message.chat.id]['price'] = price
    admin_inputs.append(price)
    bot.send_message(message.chat.id, 'Введите наименование товара')
    bot.register_next_step_handler(message, admin_name, key)


def admin_name(message, key):
    name = message.text.strip()
    user_inputs[message.chat.id]['name'] = name
    admin_inputs.append(name)
    bot.send_message(message.chat.id, 'Введите ссылку на товар')
    bot.register_next_step_handler(message, admin_photo, key)


def admin_photo(message, key):
    photo = message.text.strip()
    user_inputs[message.chat.id]['photo'] = photo
    admin_inputs.append(photo)
    bot.send_message(message.chat.id, 'Напишите "ОК" что бы продолжить')
    bot.register_next_step_handler(message, create_button, key)


@bot.message_handler(commands=['start'])
def start(message):
    user_inputs[message.chat.id] = {}
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Регистрация",callback_data='Registration'))
    bot.send_message(message.chat.id, 'Приветсвуем тебя дорогой пользователь', reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == 'Registration')
def registration_callback(call):
    message= call.message
    bot.send_message(message.chat.id, 'Добро пожаловать в наш магазин детской одежды!')
    bot.send_message(message.chat.id,'Давайте зарегистрируем вас, чтобы подобрать подходящую одежду и обувь. Введите ваш рост:')
    bot.register_next_step_handler(message, user_height)



@bot.message_handler(commands=['menu'])
def menu(message):
    key = ''
    create_button(message, key)


def user_height(message):
    height = message.text.strip()
    user_inputs[message.chat.id]['height'] = height
    bot.send_message(message.chat.id, 'Введите размер обуви:')
    bot.register_next_step_handler(message, user_shoe_size)


def create_button(message, key):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Перейти на наш сайт", url='https://vk.com/bonitodeti'))
    markup.add(types.InlineKeyboardButton('Обувь', callback_data=Const.shoes + key))
    markup.add(types.InlineKeyboardButton('Головные уборы', callback_data=Const.hat + key))
    markup.add(types.InlineKeyboardButton('Одежда', callback_data=Const.cloth + key))

    bot.send_message(message.chat.id, 'Выберите категорию:', reply_markup=markup)


def user_shoe_size(message):
    try:
        shoe_size = message.text.strip()
        user_inputs[message.chat.id]['shoe_size'] = shoe_size

        bot.send_message(message.chat.id,'Спасибо за регистрацию! Теперь мы можем подобрать для вас подходящую одежду и обувь.')

        user_data = user_inputs[message.chat.id]
        response = f"Ваши данные:\nРост: {user_data['height']}\nРазмер обуви: {user_data['shoe_size']}"
        bot.send_message(message.chat.id, response)
        # работа с кнопками
        key = ''
        create_button(message, key)

    except ValueError:
        bot.send_message(message.chat.id, 'Некорректный ввод. Пожалуйста, введите только числовое значение размера')
        bot.register_next_step_handler(message, user_shoe_size)


def delete_available(call_data,id,message):
    book = openpyxl.load_workbook(filename='data.xlsx')
    sheet = None
    key = '_delete'
    if call_data == Const.shoes+key:
        sheet:worksheet = book.worksheets[0]

    elif call_data == Const.cloth+key:
        sheet:worksheet = book.worksheets[1]
    elif call_data == Const.hat+key:
        sheet:worksheet = book.worksheets[2]
    delete_item(sheet,id,book,message)


def set_available(call_data,call):
    book = openpyxl.load_workbook(filename='data.xlsx')
    sheet = None
    available = []
    key ='_admin'

    if call_data == Const.shoes+key:
        sheet:worksheet = book.worksheets[0]

    elif call_data == Const.cloth+key:
        sheet:worksheet = book.worksheets[1]
    elif call_data == Const.hat+key:
        sheet:worksheet = book.worksheets[2]

    user_data = user_inputs[call.message.chat.id]

    sheet.insert_rows(0)
    sheet['A1'].value  = admin_inputs[2]
    sheet['B1'].value  = admin_inputs[0]
    sheet['C1'].value  = admin_inputs[3]
    sheet['D1'].value  = admin_inputs[1]
    sheet['E1'].value = id_generator(sheet)
    book.save(filename='data.xlsx')
    bot.send_message(call.message.chat.id,'Данные внесены')


def id_generator(sheet):
    while True:
        id = random.randint(10000, 99999)

        # Проверяем, что идентификатор не существует в столбце E таблицы
        column_e = sheet['E']
        for cell in column_e:
            if cell.value == id:
                id_generator(sheet)
        else:
            return id
def delete_item(sheet,id,workbook,message):
    column_e = sheet['E']
    for cell in column_e:
        if cell.value == id:
            print('Идентификатор найден')
            target_row = cell.row
            sheet.delete_rows(target_row)
            workbook.save("data.xlsx")

            break
    bot.send_message(message.chat.id,'Данные удалены')

def get_available(user_size, call_data):
    workbook = openpyxl.load_workbook(filename='data.xlsx')
    sheet = None
    available = []

    if call_data == Const.shoes:
        sheet = workbook['Обувь']
    elif call_data == Const.cloth:
        sheet = workbook['Одежда']
    elif call_data == Const.hat:
        sheet = workbook['Головные уборы']

    for row in sheet.iter_rows(min_row=1, values_only=True):
        brand = row[0]
        size = row[1]
        photo_url = row[2]
        price = row[3]
        if call_data == 'shoes' or call_data == 'clothes':
            if str(size) == user_size:
                available.append((brand, size, photo_url, price))
        else:
            available.append((brand, size, photo_url, price))

    workbook.close()
    return available


@bot.callback_query_handler(func=lambda call: call.data == 'shoes_admin')
def shoes_callback_admin(call):
    call_data = 'shoes_admin'
    set_available(call_data,call)

@bot.callback_query_handler(func=lambda call: call.data == 'shoes_delete')
def shoes_callback_delete(call):
    message = call.message
    call_data = 'shoes_delete'
    bot.send_message(call.message.chat.id, 'Введите идентификатор товара')
    bot.register_next_step_handler(message,get_message_id,call_data)
@bot.callback_query_handler(func=lambda call: call.data == 'clothes_delete')
def clothes_callback_delete(call):
    message = call.message
    call_data = 'clothes_delete'
    bot.send_message(call.message.chat.id, 'Введите идентификатор товара')
    bot.register_next_step_handler(message,get_message_id,call_data)

@bot.callback_query_handler(func=lambda call: call.data == 'hats_delete')
def hats_callback_delete(call):
        message = call.message
        call_data = 'hats_delete'
        bot.send_message(call.message.chat.id, 'Введите идентификатор товара')
        bot.register_next_step_handler(message, get_message_id,call_data)

def get_message_id(message,call_data):
    id = message.text.strip()
    delete_available(call_data,int(id),message)



@bot.callback_query_handler(func=lambda call: call.data == Const.shoes)
def shoes_callback(call):
    user_data = user_inputs[call.message.chat.id]
    shoe_size = user_data['shoe_size']
    call_data = 'shoes'
    available_shoes = get_available(shoe_size, call_data)

    if available_shoes:
        response = f"Доступная обувь размера {shoe_size}:\n"
        bot.send_message(call.message.chat.id, response)
        for shoe in available_shoes:
            brand = shoe[0]
            size = shoe[1]
            photo_url = shoe[2]
            price = shoe[3]

            bot.send_photo(call.message.chat.id, photo_url, caption=f"Бренд: {brand}\nРазмер: {size}\n Цена: {price}")
    else:
        response = "Извините, но нет доступной обуви в вашем размере."
        bot.send_message(call.message.chat.id, response)


@bot.callback_query_handler(func=lambda call: call.data == 'clothes_admin')
def clothes_callback_admin(call):
    call_data = 'clothes_admin'
    set_available(call_data,call)



@bot.callback_query_handler(func=lambda call: call.data == Const.cloth)
def clothes_callback(call):
    user_data = user_inputs[call.message.chat.id]
    clothes_size = user_data['height']
    call_data = 'clothes'
    current_size = 32 + (int(clothes_size) - 125) // 6
    available_clothes = get_available(str(current_size), call_data)

    if available_clothes:
        response = f"Доступная одежда размера {clothes_size}:\n"
        bot.send_message(call.message.chat.id, response)
        for clothes in available_clothes:
            brand = clothes[0]
            size = clothes[1]
            photo_url = clothes[2]
            price = clothes[3]

            bot.send_photo(call.message.chat.id, photo_url, caption=f"Бренд: {brand}\nРазмер: {size}\n Цена: {price}")
    else:
        response = "Извините, но нет доступной одежды в вашем размере."
        bot.send_message(call.message.chat.id, response)


@bot.callback_query_handler(func=lambda call: call.data == 'hats_admin')
def hats_callback_admin(call):
    call_data = 'hats_admin'
    set_available(call_data,call)


@bot.callback_query_handler(func=lambda call: call.data == Const.hat)
def hats_callback(call):
    call_data = 'hats'
    available = get_available(0, call_data)

    if available:
        response = f"Доступные головные уборы:\n Цена: {price}"
        bot.send_message(call.message.chat.id, response)
        for hat in available:
            photo_url = hat[2]
            price = hat[3]
            bot.send_photo(call.message.chat.id, photo_url, caption=hat[2])
    else:
        response = "Извините, но нет доступных головных уборов."
        bot.send_message(call.message.chat.id, response)


bot.polling(none_stop=True)